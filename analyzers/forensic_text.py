"""
Forensic Text Analyzer
Extracts forensic text patterns: emails, URLs, UNC paths, IP addresses, 
hidden text, and temporal anomalies.
"""

import re
import datetime
from utils.helpers import NS, log_info, log_warning, log_danger, log_success
from xml.etree import ElementTree as ET

class ForensicTextAnalyzer:
    def __init__(self, loader):
        self.loader = loader
        self.emails = set()
        self.unc_paths = set()
        self.ip_addresses = set()
        self.hidden_text = []
        self.future_timestamps = []
        
    def run(self):
        print("\n--- Forensic Text Pattern Analysis ---")
        self._extract_emails()
        self._extract_unc_paths()
        self._extract_ip_addresses()
        # Note: URLs are already checked by HyperlinkAnalyzer
        self._detect_hidden_text()
        self._detect_temporal_anomalies()
        
    def _extract_emails(self):
        """Extract email addresses from document content and metadata."""
        email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
        
        def extract_text_from_tree(tree):
            """Helper to safely extract text from Element or ElementTree."""
            if tree is None:
                return ""
            root = tree.getroot() if hasattr(tree, 'getroot') else tree
            return ET.tostring(root, encoding='unicode', method='text')
        
        # Check document content
        if self.loader.file_type == 'docx':
            tree = self.loader.get_xml_tree('word/document.xml')
            if tree:
                text_content = extract_text_from_tree(tree)
                emails = re.findall(email_pattern, text_content)
                self.emails.update(emails)
                
            # Check headers/footers
            for part in ['header', 'footer']:
                for xml_file in self.loader.zip_ref.namelist():
                    if xml_file.startswith(f'word/{part}') and xml_file.endswith('.xml'):
                        tree = self.loader.get_xml_tree(xml_file)
                        if tree:
                            text_content = extract_text_from_tree(tree)
                            emails = re.findall(email_pattern, text_content)
                            self.emails.update(emails)
                            
        elif self.loader.file_type == 'xlsx':
            from openpyxl import load_workbook
            try:
                wb = load_workbook(self.loader.filepath, read_only=True, data_only=True)
                for sheet in wb.worksheets:
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value and isinstance(cell.value, str):
                                emails = re.findall(email_pattern, cell.value)
                                self.emails.update(emails)
                wb.close()
            except Exception as e:
                pass
                
        elif self.loader.file_type in ['odt', 'ods', 'odp']:
            tree = self.loader.get_xml_tree('content.xml')
            if tree:
                text_content = extract_text_from_tree(tree)
                emails = re.findall(email_pattern, text_content)
                self.emails.update(emails)
        
        # Check metadata
        for meta_file in ['docProps/core.xml', 'docProps/app.xml', 'meta.xml']:
            tree = self.loader.get_xml_tree(meta_file)
            if tree:
                meta_text = extract_text_from_tree(tree)
                emails = re.findall(email_pattern, meta_text)
                self.emails.update(emails)
        
        if self.emails:
            log_warning(f"Found {len(self.emails)} email address(es):")
            for email in sorted(self.emails)[:10]:
                print(f"  → {email}")
            if len(self.emails) > 10:
                print(f"  ... and {len(self.emails) - 10} more")
        else:
            log_success("No email addresses found.")
    
    def _extract_unc_paths(self):
        """Extract UNC paths that might reveal internal network structure."""
        unc_pattern = r'\\\\[a-zA-Z0-9_\-\.]+\\[a-zA-Z0-9_\-\.\$\\]+'
        
        def extract_text_from_tree(tree):
            """Helper to safely extract text from Element or ElementTree."""
            if tree is None:
                return ""
            root = tree.getroot() if hasattr(tree, 'getroot') else tree
            return ET.tostring(root, encoding='unicode', method='text')
        
        # Check document content
        if self.loader.file_type == 'docx':
            tree = self.loader.get_xml_tree('word/document.xml')
            if tree:
                text_content = extract_text_from_tree(tree)
                paths = re.findall(unc_pattern, text_content)
                self.unc_paths.update(paths)
                
        elif self.loader.file_type == 'xlsx':
            from openpyxl import load_workbook
            try:
                wb = load_workbook(self.loader.filepath, read_only=True, data_only=True)
                for sheet in wb.worksheets:
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value and isinstance(cell.value, str):
                                paths = re.findall(unc_pattern, cell.value)
                                self.unc_paths.update(paths)
                wb.close()
            except Exception as e:
                pass
                
        elif self.loader.file_type in ['odt', 'ods', 'odp']:
            tree = self.loader.get_xml_tree('content.xml')
            if tree:
                text_content = extract_text_from_tree(tree)
                paths = re.findall(unc_pattern, text_content)
                self.unc_paths.update(paths)
        
        # Check relationships for external links
        for rel_file in self.loader.zip_ref.namelist():
            if '_rels' in rel_file and rel_file.endswith('.rels'):
                tree = self.loader.get_xml_tree(rel_file)
                if tree:
                    rel_text = extract_text_from_tree(tree)
                    paths = re.findall(unc_pattern, rel_text)
                    self.unc_paths.update(paths)
        
        if self.unc_paths:
            log_danger(f"Found {len(self.unc_paths)} UNC path(s) - reveals internal network:")
            for path in sorted(self.unc_paths)[:10]:
                print(f"  → {path}")
            if len(self.unc_paths) > 10:
                print(f"  ... and {len(self.unc_paths) - 10} more")
        else:
            log_success("No UNC paths found.")
    
    def _extract_ip_addresses(self):
        """Extract IP addresses from document content."""
        ip_pattern = r'\b(?:\d{1,3}\.){3}\d{1,3}\b'
        
        def extract_text_from_tree(tree):
            """Helper to safely extract text from Element or ElementTree."""
            if tree is None:
                return ""
            root = tree.getroot() if hasattr(tree, 'getroot') else tree
            return ET.tostring(root, encoding='unicode', method='text')
        
        # Check document content
        if self.loader.file_type == 'docx':
            tree = self.loader.get_xml_tree('word/document.xml')
            if tree:
                text_content = extract_text_from_tree(tree)
                ips = re.findall(ip_pattern, text_content)
                # Validate IPs
                for ip in ips:
                    parts = ip.split('.')
                    if all(0 <= int(p) <= 255 for p in parts):
                        self.ip_addresses.add(ip)
                        
        elif self.loader.file_type == 'xlsx':
            from openpyxl import load_workbook
            try:
                wb = load_workbook(self.loader.filepath, read_only=True, data_only=True)
                for sheet in wb.worksheets:
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value and isinstance(cell.value, str):
                                ips = re.findall(ip_pattern, cell.value)
                                for ip in ips:
                                    parts = ip.split('.')
                                    if all(0 <= int(p) <= 255 for p in parts):
                                        self.ip_addresses.add(ip)
                wb.close()
            except Exception as e:
                pass
                
        elif self.loader.file_type in ['odt', 'ods', 'odp']:
            tree = self.loader.get_xml_tree('content.xml')
            if tree:
                text_content = extract_text_from_tree(tree)
                ips = re.findall(ip_pattern, text_content)
                for ip in ips:
                    parts = ip.split('.')
                    if all(0 <= int(p) <= 255 for p in parts):
                        self.ip_addresses.add(ip)
        
        if self.ip_addresses:
            log_warning(f"Found {len(self.ip_addresses)} IP address(es):")
            for ip in sorted(self.ip_addresses):
                print(f"  → {ip}")
        else:
            log_success("No IP addresses found.")
    
    def _detect_hidden_text(self):
        """Detect potentially hidden text (white-on-white, zero-size, etc.)."""
        if self.loader.file_type != 'docx':
            return
            
        tree = self.loader.get_xml_tree('word/document.xml')
        if not tree:
            return
        
        # Check for white text
        white_colors = ['FFFFFF', 'ffffff', 'white']
        runs = tree.xpath('//w:r', namespaces=NS)
        
        for run in runs:
            text_elem = run.find('.//w:t', NS)
            if text_elem is not None and text_elem.text:
                # Check text color
                color_elem = run.find('.//w:color', NS)
                if color_elem is not None:
                    color = color_elem.get(f"{{{NS['w']}}}val", '')
                    if color in white_colors:
                        text = text_elem.text.strip()
                        if text:
                            self.hidden_text.append({
                                'type': 'White text',
                                'content': text[:50] + '...' if len(text) > 50 else text
                            })
                
                # Check for vanish (hidden) property
                vanish = run.find('.//w:vanish', NS)
                if vanish is not None:
                    val = vanish.get(f"{{{NS['w']}}}val", '1')
                    if val in ['1', 'true']:
                        text = text_elem.text.strip()
                        if text:
                            self.hidden_text.append({
                                'type': 'Hidden (vanish)',
                                'content': text[:50] + '...' if len(text) > 50 else text
                            })
                
                # Check for very small text (< 1pt)
                sz_elem = run.find('.//w:sz', NS)
                if sz_elem is not None:
                    size = sz_elem.get(f"{{{NS['w']}}}val", '')
                    try:
                        if int(size) < 2:  # Half-points (1pt = 2)
                            text = text_elem.text.strip()
                            if text:
                                self.hidden_text.append({
                                    'type': f'Tiny text ({int(size)/2}pt)',
                                    'content': text[:50] + '...' if len(text) > 50 else text
                                })
                    except:
                        pass
        
        if self.hidden_text:
            log_danger(f"Found {len(self.hidden_text)} potentially hidden text element(s):")
            for item in self.hidden_text[:5]:
                print(f"  [{item['type']}]: \"{item['content']}\"")
            if len(self.hidden_text) > 5:
                print(f"  ... and {len(self.hidden_text) - 5} more")
        else:
            log_success("No hidden text detected.")
    
    def _detect_temporal_anomalies(self):
        """Detect future-dated timestamps and other temporal anomalies."""
        now = datetime.datetime.now()
        
        # Check metadata timestamps
        timestamps = []
        
        if self.loader.file_type in ['docx', 'xlsx', 'pptx']:
            tree = self.loader.get_xml_tree('docProps/core.xml')
            if tree:
                ns = {'dcterms': 'http://purl.org/dc/terms/'}
                for field in ['created', 'modified']:
                    elems = tree.xpath(f'//dcterms:{field}', namespaces=ns)
                    if elems and elems[0].text:
                        try:
                            timestamp = datetime.datetime.fromisoformat(elems[0].text.rstrip('Z'))
                            if timestamp > now:
                                self.future_timestamps.append({
                                    'field': field,
                                    'value': elems[0].text,
                                    'days_ahead': (timestamp - now).days
                                })
                        except:
                            pass
                            
        elif self.loader.file_type in ['odt', 'ods', 'odp']:
            tree = self.loader.get_xml_tree('meta.xml')
            if tree:
                ns = {'meta': 'urn:oasis:names:tc:opendocument:xmlns:meta:1.0',
                      'dc': 'http://purl.org/dc/elements/1.1/'}
                for field, xpath in [('created', '//meta:creation-date'), 
                                    ('modified', '//dc:date')]:
                    elems = tree.xpath(xpath, namespaces=ns)
                    if elems and elems[0].text:
                        try:
                            timestamp = datetime.datetime.fromisoformat(elems[0].text.rstrip('Z'))
                            if timestamp > now:
                                self.future_timestamps.append({
                                    'field': field,
                                    'value': elems[0].text,
                                    'days_ahead': (timestamp - now).days
                                })
                        except:
                            pass
        
        if self.future_timestamps:
            log_danger("Found future-dated timestamp(s) - SUSPICIOUS:")
            for ts in self.future_timestamps:
                print(f"  {ts['field']}: {ts['value']} ({ts['days_ahead']} days in future)")
        else:
            log_success("No temporal anomalies detected.")
