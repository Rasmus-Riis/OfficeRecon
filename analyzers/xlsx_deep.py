"""
XLSX Deep Analyzer - Forensic Analysis for Excel Files
Extracts metadata, hidden content, comments, macros, and structural anomalies.
"""
from utils.helpers import NS, log_info, log_warning, log_success, log_danger
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class XLSXDeepAnalyzer:
    def __init__(self, loader):
        self.loader = loader
        self.workbook = None
        
    def run(self):
        print("\n--- Excel Specific Forensics ---")
        
        if not OPENPYXL_AVAILABLE:
            log_warning("openpyxl not available. Install it for deep XLSX analysis.")
            return
        
        try:
            # Load workbook with data_only=False to preserve formulas
            self.workbook = openpyxl.load_workbook(self.loader.filepath, data_only=False, keep_vba=True)
            
            self._analyze_metadata()
            self._scan_sheets()
            self._scan_hidden_content()
            self._scan_comments()
            self._scan_defined_names()
            self._scan_external_links()
            self._scan_data_validation()
            self._scan_formulas()
            self._check_protection()
            self._check_macros()
            self._scan_custom_properties()
            
        except Exception as e:
            log_danger(f"Error loading XLSX file: {e}")

    def _analyze_metadata(self):
        """Extract core metadata from workbook properties."""
        print(f"\n{'[XLSX Metadata]':<25}")
        
        props = self.workbook.properties
        
        metadata = {
            'Title': props.title,
            'Author': props.creator,
            'Last Modified By': props.lastModifiedBy,
            'Created': props.created,
            'Modified': props.modified,
            'Company': props.company,
            'Description': props.description,
            'Subject': props.subject,
            'Keywords': props.keywords,
            'Category': props.category,
            'Comments': props.comments,
            'Version': props.version,
            'Revision': props.revision,
            'Content Status': props.contentStatus,
        }
        
        for label, value in metadata.items():
            if value:
                print(f"  {label:<20}: {value}")

    def _scan_sheets(self):
        """Analyze all sheets including visibility status."""
        print(f"\n{'[Sheet Analysis]':<25}")
        
        sheet_names = self.workbook.sheetnames
        print(f"  Total Sheets: {len(sheet_names)}")
        
        visible_sheets = []
        hidden_sheets = []
        very_hidden_sheets = []
        
        for sheet_name in sheet_names:
            sheet = self.workbook[sheet_name]
            if sheet.sheet_state == 'visible':
                visible_sheets.append(sheet_name)
            elif sheet.sheet_state == 'hidden':
                hidden_sheets.append(sheet_name)
            elif sheet.sheet_state == 'veryHidden':
                very_hidden_sheets.append(sheet_name)
        
        print(f"  Visible: {len(visible_sheets)}")
        if visible_sheets:
            print(f"    -> {', '.join(visible_sheets[:5])}")
            if len(visible_sheets) > 5:
                print(f"    -> ... and {len(visible_sheets) - 5} more")
        
        if hidden_sheets:
            log_warning(f"Hidden Sheets ({len(hidden_sheets)}): {', '.join(hidden_sheets)}")
        
        if very_hidden_sheets:
            log_danger(f"VERY HIDDEN Sheets ({len(very_hidden_sheets)}): {', '.join(very_hidden_sheets)}")

    def _scan_hidden_content(self):
        """Scan for hidden rows and columns."""
        print(f"\n{'[Hidden Rows/Columns]':<25}")
        
        findings = []
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            hidden_rows = []
            hidden_cols = []
            
            # Check hidden rows (sample first 1000 rows)
            for row in range(1, min(1001, sheet.max_row + 1)):
                if sheet.row_dimensions[row].hidden:
                    hidden_rows.append(row)
            
            # Check hidden columns
            for col in range(1, sheet.max_column + 1):
                col_letter = get_column_letter(col)
                if sheet.column_dimensions[col_letter].hidden:
                    hidden_cols.append(col_letter)
            
            if hidden_rows or hidden_cols:
                findings.append(f"  Sheet '{sheet_name}':")
                if hidden_rows:
                    log_warning(f"    Hidden Rows: {len(hidden_rows)} (e.g., {hidden_rows[:5]})")
                if hidden_cols:
                    log_warning(f"    Hidden Columns: {', '.join(hidden_cols[:10])}")
        
        if not findings:
            log_success("No hidden rows or columns detected.")

    def _scan_comments(self):
        """Extract all comments with author and location."""
        print(f"\n{'[Comments & Annotations]':<25}")
        
        total_comments = 0
        authors = set()
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            sheet_comments = []
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.comment:
                        total_comments += 1
                        author = cell.comment.author if cell.comment.author else "Unknown"
                        authors.add(author)
                        text = cell.comment.text[:80] + "..." if len(cell.comment.text) > 80 else cell.comment.text
                        sheet_comments.append((cell.coordinate, author, text))
            
            if sheet_comments:
                log_warning(f"Sheet '{sheet_name}' has {len(sheet_comments)} comments:")
                for coord, author, text in sheet_comments[:3]:
                    print(f"    [{coord}] {author}: \"{text}\"")
                if len(sheet_comments) > 3:
                    print(f"    ... and {len(sheet_comments) - 3} more")
        
        if total_comments > 0:
            print(f"\n  Total Comments: {total_comments}")
            print(f"  Comment Authors: {', '.join(authors)}")
        else:
            log_success("No comments found.")

    def _scan_defined_names(self):
        """Scan for defined names (which can hide data or formulas)."""
        print(f"\n{'[Defined Names]':<25}")
        
        if not self.workbook.defined_names:
            log_success("No defined names found.")
            return
        
        log_info(f"Found {len(self.workbook.defined_names.definedName)} defined names")
        
        suspicious = []
        for name in self.workbook.defined_names.definedName:
            name_str = name.name
            dest = name.attr_text
            
            # Check for suspicious patterns
            if any(keyword in dest.lower() for keyword in ['http', 'https', 'ftp', '\\\\', 'cmd', 'powershell']):
                suspicious.append((name_str, dest))
            
            print(f"  {name_str:<20}: {dest[:60]}")
        
        if suspicious:
            log_danger(f"SUSPICIOUS defined names detected:")
            for name, dest in suspicious:
                print(f"    -> {name}: {dest}")

    def _scan_external_links(self):
        """Detect external data connections and links."""
        print(f"\n{'[External Links & Connections]':<25}")
        
        # Check for external workbook links via XML
        tree = self.loader.get_xml_tree('xl/workbook.xml')
        if tree:
            ns = {'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}
            ext_refs = tree.xpath('//externalReference', namespaces=ns)
            
            if ext_refs:
                log_warning(f"Found {len(ext_refs)} external workbook references")
        
        # Check connections
        if self.loader.file_exists('xl/connections.xml'):
            log_warning("External data connections detected (connections.xml exists)")
            conn_tree = self.loader.get_xml_tree('xl/connections.xml')
            if conn_tree:
                connections = conn_tree.xpath('//*[local-name()="connection"]')
                print(f"  Connection count: {len(connections)}")
        else:
            log_success("No external connections found.")

    def _scan_data_validation(self):
        """Check for data validation rules (can contain formulas)."""
        print(f"\n{'[Data Validation]':<25}")
        
        validation_count = 0
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            
            if hasattr(sheet, 'data_validations') and sheet.data_validations:
                validations = sheet.data_validations.dataValidation
                if validations:
                    validation_count += len(validations)
                    log_info(f"Sheet '{sheet_name}': {len(validations)} validation rules")
        
        if validation_count == 0:
            log_success("No data validation rules found.")

    def _scan_formulas(self):
        """Scan for potentially dangerous formulas."""
        print(f"\n{'[Formula Analysis]':<25}")
        
        suspicious_formulas = []
        formula_count = 0
        
        dangerous_functions = [
            'HYPERLINK', 'WEBSERVICE', 'FILTERXML', 'INDIRECT',
            'EXEC', 'CALL', 'REGISTER', 'SYSTEM'
        ]
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            
            for row in sheet.iter_rows(max_row=min(1000, sheet.max_row)):
                for cell in row:
                    if cell.data_type == 'f' and cell.value:
                        formula_count += 1
                        formula_upper = str(cell.value).upper()
                        
                        for func in dangerous_functions:
                            if func in formula_upper:
                                suspicious_formulas.append((sheet_name, cell.coordinate, cell.value))
                                break
        
        print(f"  Total Formulas: {formula_count}")
        
        if suspicious_formulas:
            log_danger(f"SUSPICIOUS FORMULAS DETECTED ({len(suspicious_formulas)}):")
            for sheet, coord, formula in suspicious_formulas[:5]:
                print(f"    [{sheet}!{coord}] {formula[:80]}")
            if len(suspicious_formulas) > 5:
                print(f"    ... and {len(suspicious_formulas) - 5} more")
        else:
            log_success("No suspicious formulas detected.")

    def _check_protection(self):
        """Check for workbook and sheet protection."""
        print(f"\n{'[Protection Status]':<25}")
        
        if self.workbook.security:
            log_warning("Workbook structure is PROTECTED")
            if hasattr(self.workbook.security, 'workbookPassword'):
                print("  -> Password-protected structure detected")
        
        protected_sheets = []
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            if sheet.protection.sheet:
                protected_sheets.append(sheet_name)
        
        if protected_sheets:
            log_warning(f"Protected Sheets ({len(protected_sheets)}): {', '.join(protected_sheets)}")
        else:
            log_success("No sheet protection detected.")

    def _check_macros(self):
        """Check for VBA macros in XLSX (rare but possible)."""
        print(f"\n{'[Macro Detection]':<25}")
        
        # Check if VBA project exists
        if self.loader.file_exists('xl/vbaProject.bin'):
            log_danger("VBA MACROS DETECTED! This file contains executable code.")
            print("  -> File should be analyzed as XLSM (macro-enabled)")
        else:
            log_success("No VBA macros detected.")

    def _scan_custom_properties(self):
        """Extract custom document properties."""
        tree = self.loader.get_xml_tree('docProps/custom.xml')
        if not tree:
            return
        
        print(f"\n{'[Custom Properties]':<25}")
        
        ns = {'cp': 'http://schemas.openxmlformats.org/officeDocument/2006/custom-properties',
              'vt': 'http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes'}
        
        props = tree.xpath('//cp:property', namespaces=ns)
        
        if props:
            for prop in props:
                name = prop.get('name', 'Unknown')
                # Try to get the value from various possible child elements
                value_elem = prop.xpath('./vt:lpwstr | ./vt:i4 | ./vt:bool | ./vt:filetime', namespaces=ns)
                value = value_elem[0].text if value_elem and value_elem[0].text else 'N/A'
                print(f"  {name:<20}: {value}")
        else:
            log_success("No custom properties found.")
