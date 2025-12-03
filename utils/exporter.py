import os
import re
import platform
import subprocess
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# --- FIX: Define Regex locally to avoid ImportErrors on newer OpenPyXL versions ---
# Matches characters that are illegal in XML (Excel) files
# ASCII Control chars (0-31) EXCEPT Tab (9), Newline (10), Carriage Return (13)
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

def clean_text(text):
    """Removes ANSI color codes and illegal XML characters."""
    if not isinstance(text, str): return text
    
    # 1. Remove ANSI Escape codes (e.g. colors from ExifTool)
    ansi_escape = re.compile(r'\x1B(?:[@-Z\\-_]|\[[0-?]*[ -/]*[@-~])')
    text = ansi_escape.sub('', text)
    
    # 2. Remove Excel Illegal Characters
    text = ILLEGAL_CHARACTERS_RE.sub('', text)
    
    return text

def export_to_excel(table_data, columns):
    if not table_data:
        messagebox.showwarning("Export", "No data to export.")
        return

    path = filedialog.asksaveasfilename(defaultextension=".xlsx", 
                                        filetypes=[("Excel Workbook", "*.xlsx")])
    if not path: return

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Forensic Report"

        # Headers
        headers = [c["label"] for c in columns]
        ws.append(headers)

        # Rows
        for row in table_data:
            row_values = []
            for c in columns:
                key = c["key"]
                # Use the FULL raw text for the export (not the "View Report" placeholder)
                val = row.get('deep_output_raw', '') if key == 'deep_output' else row.get(key, "")
                row_values.append(clean_text(str(val)))
            ws.append(row_values)

        # Formatting
        ws.freeze_panes = 'A2'
        for i, col in enumerate(ws.columns, 1):
            max_len = 0
            column_letter = get_column_letter(i)
            for cell in col:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                try:
                    if cell.value:
                        # Estimate width based on first line length
                        line_len = len(str(cell.value).split('\n')[0])
                        if line_len > max_len: max_len = line_len
                except: pass
            
            # Cap width at 80 to prevent massive columns
            ws.column_dimensions[column_letter].width = min(max_len + 2, 80)

        wb.save(path)

        if messagebox.askyesno("Export Successful", "Data exported successfully!\n\nOpen containing folder?"):
            _open_folder(os.path.dirname(path))

    except Exception as e:
        messagebox.showerror("Export Error", f"Failed to save file:\n{str(e)}")

def _open_folder(path):
    try:
        if platform.system() == "Windows": os.startfile(path)
        elif platform.system() == "Darwin": subprocess.run(["open", path])
        else: subprocess.run(["xdg-open", path])
    except Exception as e: print(e)